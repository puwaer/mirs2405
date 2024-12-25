import os
import numpy as np
import pandas as pd
import copy
from openpyxl import Workbook, load_workbook
import pdb

from color import Color


# Excelファイルに出力
def Excel_output(data, start_row, start_col):
    data2 = copy.deepcopy(data)
    for i in range (len(data2)):
         data_i = data2[i]
         if data_i[6] != True:
             continue
         data_i[7] = np.rad2deg(data_i[7])
         data_i[8] = np.rad2deg(data_i[8])
         data_i[9] = np.rad2deg(data_i[9])

    name = "kinematics_results.xlsx"
    excel_file = os.path.join("Excel", name)
    if not os.path.exists("Excel"):
        os.makedirs("Excel")

    # データをDataFrameに変換
    df = pd.DataFrame (data2, columns = ["L_0", "L_1", "L_2", "x_target", "z_target", "到達可能か？", "θ1 (度)", "θ2 (度)", "θ3 (度)", "誤差"])

    # Excelに書き込み
    # Excelファイルを開く
    try:
        wb = load_workbook (excel_file)
    except FileNotFoundError:
        wb = Workbook()

    wb = Workbook()
    ws = wb.active

    # ヘッダーの書き込み
    for c_idx, header in enumerate (df.columns, start = start_col):
        ws.cell (row = start_row, column = c_idx, value = header)

    # データの書き込み
    for r_idx, row in enumerate (df.itertuples(index = False), start = start_row + 1):
        for c_idx, value in enumerate (row, start = start_col):
            ws.cell (row = r_idx, column = c_idx, value = value)

    # 保存
    wb.save (excel_file)
    print (Color.YELLOW, f"データが {excel_file} に保存されました。", Color.RESET, sep = '')