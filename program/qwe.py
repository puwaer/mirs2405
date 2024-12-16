import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

# 入力データ
L = [0.5, 0.5, 0.1]  # 各リンクの長さ
x_target, z_target = 0.6, 0.7  # 目標座標
L_target = np.sqrt(x_target**2 + z_target**2)

start_row, start_col = 2, 2 # Excelファイルへ書き込み始めるセル

data = []

def truncate(x_target, z_target):
    factor = 10 ** 1
    x_target = int(x_target * factor) / factor
    z_target = int(z_target * factor) / factor
    return (x_target, z_target)

def fig_plot(θ_1, θ_2, θ_3):
    x_1 = L[0] * np.cos(θ_1)
    z_1 = L[0] * np.sin(θ_1)
    x_2 = x_1 + L[1] * np.cos(θ_1 + θ_2)
    z_2 = z_1 + L[1] * np.sin(θ_1 + θ_2)
    x_3 = x_2 + L[2] * np.cos(θ_1 + θ_2 + θ_3)
    z_3 = z_2 + L[2] * np.sin(θ_1 + θ_2 + θ_3)

    X = [0, x_1, x_2, x_3]
    Z = [0, z_1, z_2, z_3]

    color = ["blue", "green", "red"]

    # アーム全体を描画
    for i in range (0, 3, 1):
        pX = [X[i], X[i+1]]
        pZ = [Z[i], Z[i+1]]
        plt.plot(pX, pZ, label='Arm Configuration', color = color[i])
        # plt.plot(X[i], Z[i], marker='o', color = 'black')
    plt.plot(X[3], Z[3], marker='o', color = 'yellow')
    plt.axhline(0, color="black", linewidth=0.8)  # X軸
    plt.axvline(0, color="black", linewidth=0.8)  # Z軸
    plt.xlabel("X-axis")
    plt.ylabel("Z-axis")
    plt.title("Arm Configuration at Current Target")
    plt.grid(True)

   
def keisan(x_target, z_target):
    x_target, z_target = truncate(x_target, z_target)

    x_2 = x_target - L[2] * np.cos(-np.pi / 2)
    z_2 = z_target - L[2] * np.sin(-np.pi / 2)

    # 到達可能か？
    u = x_2**2 + z_2**2
    l = np.sqrt(u)

    if L[0] + L[1] < l:
        print(f"目標 ({x_target:.1f}, {z_target}) は到達不可能です。")

        # データをリストに追加
        data.append([L[0], L[1], L[2], x_target, z_target, "False", "None", "None", "None", "None"])

    cosθ_2 = (u - L[0]**2 - L[1]**2 ) / (2 * L[0] * L[1] )
    # クリップして範囲を[-1, 1]に制限
    cosθ_2 = np.clip(cosθ_2, -1.0, 1.0)
    θ_2 = np.arccos(cosθ_2)# - np.pi/2

    φ_2 = np.arctan2(x_2, z_2 )
    cosθ_1u = np.clip((u + L[0]**2 - L[1]**2) / (2 * L[0] * l), -1.0, 1.0)
    θ_1 = φ_2 + np.arccos(cosθ_1u)
    θ_3 = -np.pi/2 - θ_1 - θ_2

    x_reverse = L[0] * np.cos(θ_1) + L[1] * np.cos(θ_1 + θ_2) + L[2] * np.cos(θ_1 + θ_2 + θ_3)
    z_reverse = L[0] * np.sin(θ_1) + L[1] * np.sin(θ_1 + θ_2) + L[2] * np.sin(θ_1 + θ_2 + θ_3)

    # 目標値との誤差
    error_x = x_target - x_reverse
    error_z = z_target - z_reverse
    error = np.sqrt(error_x**2 + error_z**2)

    # degreeに変換
    θ_1_deg = np.degrees(θ_1)
    θ_2_deg = np.degrees(θ_2)
    θ_3_deg = np.degrees(θ_3)
    
    print(f"cosθ_2 = {cosθ_2}, clipped = {np.clip(cosθ_2, -1.0, 1.0)}")
    print(f"cosθ_1 = {cosθ_1u}, clipped = {np.clip(cosθ_1u, -1.0, 1.0)}")

    print(f"目標 ({x_target:.1f}, {z_target}): θ_1 = {θ_1_deg:.2f}, θ_2 = {θ_2_deg:.2f}, θ_3 = {θ_3_deg:.2f}, error = {error:.2f}")


    # データをリストに追加
    data.append ([L[0], L[1], L[2], x_target, z_target, "True", θ_1_deg, θ_2_deg, θ_3_deg, error])

    fig_plot(θ_1, θ_2, θ_3)

    return(θ_1, θ_2, θ_3)

# θ_1, θ_2, θ_3 = keisan(x_target, z_target)

#"""
for z_target in np.arange(0.0, 5.1, 0.1):
    for x_target in np.arange(0.0, 5.1, 0.1):
        x_target, z_target = truncate(x_target, z_target)

        x_2 = x_target - L[2] * np.cos(-np.pi / 2)
        z_2 = z_target - L[2] * np.sin(-np.pi / 2)

        # 到達可能か？
        u = x_2**2 + z_2**2
        l = np.sqrt(u)

        if L[0] + L[1] < l:
          print(f"目標 ({x_target:.1f}, {z_target}) は到達不可能です。")

          # データをリストに追加
          data.append([L[0], L[1], L[2], x_target, z_target, "False", "None", "None", "None", "None"])
          continue

        cosθ_2 = (u - L[0]**2 - L[1]**2 ) / (2 * L[0] * L[1] )
        θ_2 = np.arccos(cosθ_2)
        if θ_2 > 0:
            θ_2 -= np.pi/2 
        elif θ_2 < 0:
            θ_2 -= np.pi/2

        φ_2 = np.arctan2(x_2, z_2 )
        θ_1 = φ_2 + np.arccos((u + L[0]**2 - L[1]**2) / (2 * L[0] * l))
        θ_3 = -np.pi/2 - θ_1 - θ_2

        x_reverse = L[0] * np.cos(θ_1) + L[1] * np.cos(θ_1 + θ_2) + L[2] * np.cos(θ_1 + θ_2 + θ_3)
        z_reverse = L[0] * np.sin(θ_1) + L[1] * np.sin(θ_1 + θ_2) + L[2] * np.sin(θ_1 + θ_2 + θ_3)

        # 目標値との誤差
        error_x = x_target - x_reverse
        error_z = z_target - z_reverse
        error = np.sqrt(error_x**2 + error_z**2)

        # degreeに変換
        θ_1_deg = np.degrees(θ_1)
        θ_2_deg = np.degrees(θ_2)
        θ_3_deg = np.degrees(θ_3)
    
        print(f"目標 ({x_target:.1f}, {z_target}): θ_1 = {θ_1_deg:.2f}, θ_2 = {θ_2_deg:.2f}, θ_3 = {θ_3_deg:.2f}, error = {error:.2f}")


        # データをリストに追加
        data.append ([L[0], L[1], L[2], x_target, z_target, "True", θ_1_deg, θ_2_deg, θ_3_deg, error])

        fig_plot(θ_1, θ_2, θ_3)

#"""

# データをDataFrameに変換
df = pd.DataFrame (data, columns = ["L_0", "L_1", "L_2", "x_target", "z_target", "到達可能か？", "θ1 (度)", "θ2 (度)", "θ3 (度)", "誤差"])

# Excelに書き込み
# Excelファイルを開く
excel_file = "kinematics_results.xlsx"
try:
    wb = load_workbook (excel_file)
except FileNotFoundError:
    wb = Workbook()

wb = Workbook()
ws = wb.active

# ヘッダーの書き込み
for c_idx, header in enumerate (df.columns, start = start_col):
    ws.cell (row = start_row, column = c_idx, value = header)

# データの書き込み
for r_idx, row in enumerate (df.itertuples(index = False), start = start_row + 1):
    for c_idx, value in enumerate (row, start = start_col):
        ws.cell (row = r_idx, column = c_idx, value = value)

# 保存
wb.save (excel_file)
print (f"データが {excel_file} に保存されました。")


plt.legend()
plt.grid(True)
plt.show ()
